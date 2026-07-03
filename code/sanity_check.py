#!/usr/bin/env python3
"""Post-run sanity checks for the MDL_new deliverable. Standalone (no shell heredoc).

Models build_excel() exactly:
  - kept order  = has appointees OR Needs_Motion_Reading is True  (keep_order policy)
  - Orders tab  = kept orders
  - Appts tab   = per-order individual/firm-appointee rows  +  derived firm rows (Appointment_ID '-F')
  - MDL id      = AUTHORITATIVE Source_File folder number (model's MDL_No used only as cross-check)
"""
import os, re, csv, json
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSONL = os.path.join(ROOT, "order_extractions.jsonl")
XLSX = os.path.join(ROOT, "order_extractions.xlsx")
GATE = os.path.join(ROOT, "order_status.csv")
CAN_ATT = os.path.join(ROOT, "canonical_attorneys.csv")
CAN_FIRM = os.path.join(ROOT, "canonical_firms.csv")


def folder_mdl(rec):
    """Authoritative MDL = leading number of the Source_File folder."""
    sf = rec.get("Source_File", "")
    m = re.match(r"^(\d{2,5})\b", sf)
    return m.group(1) if m else "?"


def appointees(rec):
    return rec.get("Appointments") or []


def keep_order(rec):
    return bool(appointees(rec)) or rec.get("Needs_Motion_Reading") is True


# ---- load jsonl ----
recs = [json.loads(l) for l in open(JSONL, encoding="utf-8") if l.strip()]

# ---- gate status ----
gate = {}
with open(GATE, newline="", encoding="utf-8") as f:
    for r in csv.DictReader(f):
        gate[r["relpath"]] = r.get("retrieve")

n_rec = len(recs)
mdls_folder = {folder_mdl(r) for r in recs}
mdls_field = {str(r.get("MDL_No")) for r in recs if r.get("MDL_No")}

gate_kept = [r for r in recs if gate.get(r.get("Source_File"), "1") == "1"]
kept = [r for r in gate_kept if keep_order(r)]                 # == Orders tab
appt_bearing = [r for r in kept if appointees(r)]
empty_needs_motion = [r for r in kept if not appointees(r)]    # kept but empty (Needs_Motion)
unresolved = [r for r in recs if r.get("Needs_Motion_Reading") is True and not appointees(r)]
docket_mm = [r for r in recs if r.get("Docket_Mismatch")]

# predicted Appointments-tab rows = raw appointee rows + derived firm rows
def norm_firm(s):
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())

raw_appt_rows = 0
derived_firm_rows = 0
for r in kept:
    ap = appointees(r)
    raw_appt_rows += len(ap)
    standalone = {norm_firm(a.get("firm")) for a in ap
                  if a.get("appointee_type") == "Firm"
                  or not ((a.get("first_name") or "").strip() or (a.get("last_name") or "").strip())}
    derived = set()
    for a in ap:
        fn = (a.get("first_name") or "").strip(); ln = (a.get("last_name") or "").strip()
        firm = (a.get("firm") or "").strip()
        if not (fn or ln) or not firm:
            continue
        nf = norm_firm(firm)
        if nf and nf not in standalone:
            derived.add(nf)
    derived_firm_rows += len(derived)
predicted_appts = raw_appt_rows + derived_firm_rows

mdls_with_kept = {folder_mdl(r) for r in kept}

print("=" * 66)
print("1) JSONL")
print(f"   records (orders extracted)        : {n_rec:,}")
print(f"   distinct MDLs (folder, authoritative): {len(mdls_folder):,}")
print(f"   distinct MDLs (model MDL_No field)   : {len(mdls_field):,}")
print(f"   gate-kept (retrieve=1)            : {len(gate_kept):,}")
print(f"   kept orders (appt OR Needs_Motion): {len(kept):,}   [Orders tab]")
print(f"      - appt-bearing                 : {len(appt_bearing):,}")
print(f"      - kept-empty (Needs_Motion)    : {len(empty_needs_motion):,}")
print()
print("2) COVERAGE (by authoritative folder MDL)")
print(f"   MDLs with >=1 kept leadership order : {len(mdls_with_kept):,}")
print(f"   MDLs with 0 kept orders             : {len(mdls_folder) - len(mdls_with_kept):,}")
print()
print("3) FLAGS")
print(f"   Needs_Motion_Reading UNRESOLVED (no appts) : {len(unresolved):,}")
print(f"   Docket_Mismatch flagged                    : {len(docket_mm):,}")
print()
print("4) WORKBOOK vs JSONL-derived (modeling build_excel exactly)")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
tab = {name: (wb[name].max_row or 1) - 1 for name in wb.sheetnames}
for name in wb.sheetnames:
    print(f"     {name:24s} rows={tab[name]:,}")
# split Appointments into individual vs derived-firm by Appointment_ID
ws = wb["Appointments"]
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
id_i = hdr.index("Appointment_ID")
dup_i = hdr.index("Possible_Duplicate_Appointment")
n_ind = n_firm = n_dup = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    aid = str(row[id_i] or "")
    if re.search(r"-F\d+$", aid):
        n_firm += 1
    else:
        n_ind += 1
    if row[dup_i] not in (None, "", False, 0, "False", "0"):
        n_dup += 1
print()

def chk(label, got, exp):
    print(f"   {label}: {got:,} vs {exp:,}  -> {'OK' if got == exp else 'MISMATCH'}")

chk("Orders tab == kept orders", tab.get("Orders", 0), len(kept))
chk("Appts tab == raw+derived", tab.get("Appointments", 0), predicted_appts)
chk("Appts individual rows == jsonl appointees", n_ind, raw_appt_rows)
chk("Appts derived-firm rows == predicted", n_firm, derived_firm_rows)
print(f"   Appointments same-date dup flags: {n_dup:,}")
print()
print("5) CANONICAL ROSTERS")
for path, label in [(CAN_ATT, "canonical_attorneys"), (CAN_FIRM, "canonical_firms")]:
    with open(path, encoding="utf-8-sig", newline="") as f:
        n = sum(1 for _ in csv.reader(f)) - 1
    print(f"   {label:22s} rows={n:,}")
print()
print("6) SPOT-CHECK 3 MDLs (order -> appointees)")
by_mdl = {}
for r in appt_bearing:
    by_mdl.setdefault(folder_mdl(r), []).append(r)
ordered = sorted(by_mdl.keys(), key=lambda x: (len(x), x))
picks = (["2275"] if "2275" in by_mdl else []) + [ordered[0], ordered[len(ordered) // 2]]
seen = set()
for m in picks:
    if m in seen:
        continue
    seen.add(m)
    rs = by_mdl[m]
    tot = sum(len(appointees(r)) for r in rs)
    print(f"   MDL {m}: {len(rs)} order(s), {tot} appointee(s)")
    for a in appointees(rs[0])[:4]:
        nm = a.get("full_name") or f"[FIRM] {a.get('firm')}"
        print(f"      - {nm} | {a.get('appointment_types')} | {a.get('plaintiff_defendant')}")
print("=" * 66)
