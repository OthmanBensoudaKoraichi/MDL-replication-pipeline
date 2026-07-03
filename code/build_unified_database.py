#!/usr/bin/env python3
"""Assemble the ONE relational database: MDLs + Orders + Appointments + Attorneys + Firms, spanning
MDL_old (gold) + MDL_new (extracted), with every appointment linked to the unified canonical
Unified_Attorney_ID / Unified_Firm_ID.

Linkage is EXACT: we re-derive the dedup clustering read-only (reusing the cached LLM adjudications),
verify it reproduces canonical_attorneys_unified.csv byte-for-byte on (ID -> Canonical_Name), and use
its key->ID maps. Nothing existing is overwritten.

Output: unified_mdl_database.xlsx
"""
import os, sys, csv, json, re
from collections import Counter, defaultdict
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "code"))
import dedup_attorneys as D
import attorney_demographics as AD

UNIFIED_JSONL = os.path.join(ROOT, "order_extractions_unified.jsonl")
ADJ_LOG = os.path.join(ROOT, "attorney_llm_adjudications_unified.csv")
ATT_ROSTER = os.path.join(ROOT, "canonical_attorneys_unified.csv")
ATT_DEMO = os.path.join(ROOT, "canonical_attorneys_unified_demographics.csv")
DEMO_CACHE = os.path.join(ROOT, "demographics_cache_unified.jsonl")
FIRM_ROSTER = os.path.join(ROOT, "canonical_firms_unified.csv")
MDL_MASTER = os.path.join(ROOT, "MDL_merged.csv")
GOLD_ORDERS = os.path.join(ROOT, "csvs_current_dataset", "Orders-Export view.csv")
GOLD_APPTS = os.path.join(ROOT, "csvs_current_dataset", "Appointments-Grid view.csv")
NEW_XLSX = os.path.join(ROOT, "order_extractions.xlsx")
OUT = os.path.join(ROOT, "unified_mdl_database.xlsx")


def load_csv(p):
    with open(p, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_xlsx(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c) if c is not None else "" for c in next(it)]
    return [dict(zip(hdr, [("" if v is None else v) for v in row])) for row in it]


# ===== 1. Re-derive attorney clustering (read-only) + build (first_tok,last)->Attorney_ID =====
print("re-deriving attorney clustering (reusing cached LLM adjudications)...")
inds, firm_mentions = D.load_mentions(UNIFIED_JSONL)
ents = D.aggregate_individuals(inds)
clusters, audit, weak, uf = D.resolve_individuals(ents)
# apply the SAME LLM merge decisions from the cached log (deterministic, no API)
adj = {}
for r in load_csv(ADJ_LOG):
    adj[frozenset((r["Name_A"], r["Name_B"]))] = str(r.get("merged")).strip().lower() in ("true", "1", "yes")
applied = 0
for i, j, _reason in weak:
    if adj.get(frozenset((D.display_name(ents[i]), D.display_name(ents[j])))):
        uf.union(i, j); applied += 1
clusters = defaultdict(list)
for i in range(len(ents)):
    clusters[uf.find(i)].append(i)
print(f"  weak pairs: {len(weak)} | cached merges applied: {applied}")

att_keymap, id_canon = {}, {}
cid = 0
for root, idxs in sorted(clusters.items(), key=lambda kv: -len(kv[1])):
    cid += 1
    aid = f"A{cid:04d}"
    names = Counter()
    for i in idxs:
        names.update(ents[i]["fulls"])
        att_keymap[(ents[i]["first_tok"], ents[i]["last"])] = aid
    id_canon[aid] = names.most_common(1)[0][0]

# verify vs the delivered roster
roster = {r["Attorney_ID"]: r["Canonical_Name"] for r in load_csv(ATT_ROSTER)}
mism = [a for a in roster if id_canon.get(a) != roster[a]]
print(f"  ATTORNEY keymap verify: {len(roster)} ids; mismatches vs roster: {len(mism)} "
      f"{'OK (exact)' if not mism else '!! '+str(mism[:5])}")

# ===== 2. Firm clustering + key->Firm_ID =====
firm_rows, firm_keymap = D.resolve_firms(firm_mentions, return_keymap=True)
firm_roster = {r["Firm_ID"]: r["Canonical_Firm"] for r in load_csv(FIRM_ROSTER)}
fid_canon = {r["Firm_ID"]: r["Canonical_Firm"] for r in firm_rows}
fmism = [f for f in firm_roster if fid_canon.get(f) != firm_roster[f]]
print(f"  FIRM keymap verify: {len(firm_roster)} ids; mismatches: {len(fmism)} "
      f"{'OK (exact)' if not fmism else '!! '+str(fmism[:5])}")


def att_id(first, last):
    return att_keymap.get((D.norm(first or ""), D.norm(last or "")), "")


def firm_id(firm):
    k = " ".join(sorted(D.firm_tokens(firm or "")))
    return firm_keymap.get(k, "")


# ===== 3. corpus MDL sets =====
gold_mdls = {(r.get("MDL_No (from Orders)") or "").strip() for r in load_csv(GOLD_APPTS)}
gold_mdls.discard("")


def nmdl(m):
    return re.sub(r"^0+(\d)", r"\1", str(m or "").strip())


gold_mdls = {nmdl(m) for m in gold_mdls}

# ===== 4. ORDERS tab (gold + new kept) =====
ORDER_COLS = ["Corpus", "Order_No", "MDL_No", "Docket_No", "MDL Type", "Date", "Judge", "Judge_Type",
              "Contested", "Applications_Solicited", "Resolve_Rule_23", "OU_Create", "OU_Terminate",
              "OU_Functions", "OU_Duties_to_Nonclients", "IRPA_Duties_to_Clients", "Limit_Nonleader_Practice",
              "Order_Types", "Rule_23", "MCL", "OU_Plaintiff", "OU_Defendant", "Appointments_Count",
              "Needs_Motion_Reading", "Provenance", "Source_File", "Notes"]
orders = []
for r in load_csv(GOLD_ORDERS):
    o = {c: r.get(c, "") for c in ORDER_COLS}
    o["Corpus"] = "old_gold"
    orders.append(o)
for r in read_xlsx(NEW_XLSX, "Orders"):
    o = {c: r.get(c, "") for c in ORDER_COLS}
    o["Corpus"] = "new_extracted"
    orders.append(o)

# ===== 5. APPOINTMENTS tab (gold + new) with relational links =====
APPT_COLS = ["Corpus", "Appointment_ID", "Order_No", "MDL_No", "MDL Type", "Last Name", "First Name",
             "First_Last_Calculated", "Firm", "Appointee Type", "Appointment Types", "Plaintiff/Defendant",
             "Appoint", "Remove", "Interim", "Unified_Attorney_ID", "Unified_Firm_ID",
             "Gold_Canonical_Name", "Possible_Duplicate_Appointment", "Provenance"]
appts = []
linked_att = linked_firm = ind_rows = 0
for r in load_csv(GOLD_APPTS):
    fn, ln, firm = r.get("First Name", ""), r.get("Last Name", ""), r.get("Firm", "")
    a = {
        "Corpus": "old_gold", "Appointment_ID": r.get("Appointment_ID", ""),
        "Order_No": r.get("Order No.") or r.get("Orders", ""), "MDL_No": nmdl(r.get("MDL_No (from Orders)", "")),
        "MDL Type": r.get("MDL Type (from Orders)", ""), "Last Name": ln, "First Name": fn,
        "First_Last_Calculated": r.get("First_Last_Calculated", ""), "Firm": firm,
        "Appointee Type": r.get("Appointee Type", ""), "Appointment Types": r.get("Appointment Types", ""),
        "Plaintiff/Defendant": r.get("Plaintiff/Defendant", ""), "Appoint": r.get("Appoint", ""),
        "Remove": r.get("Remove", ""), "Interim": r.get("Interim", ""),
        "Gold_Canonical_Name": r.get("Canonical_Name (from Attorney)", ""), "Provenance": "gold",
        "Possible_Duplicate_Appointment": "",
        "Unified_Attorney_ID": att_id(fn, ln) if (fn or ln) else "",
        "Unified_Firm_ID": firm_id(firm),
    }
    appts.append(a)
for r in read_xlsx(NEW_XLSX, "Appointments"):
    if nmdl(r.get("MDL_No", "")) in gold_mdls:
        continue   # MDL 2357 overlap -> use gold side
    fn, ln, firm = r.get("First Name", ""), r.get("Last Name", ""), r.get("Firm", "")
    a = {
        "Corpus": "new_extracted", "Appointment_ID": r.get("Appointment_ID", ""),
        "Order_No": r.get("Order_No", ""), "MDL_No": nmdl(r.get("MDL_No", "")), "MDL Type": r.get("MDL Type", ""),
        "Last Name": ln, "First Name": fn, "First_Last_Calculated": r.get("First_Last_Calculated", ""),
        "Firm": firm, "Appointee Type": r.get("Appointee Type", ""),
        "Appointment Types": r.get("Appointment Types", ""), "Plaintiff/Defendant": r.get("Plaintiff/Defendant", ""),
        "Appoint": r.get("Appoint", ""), "Remove": r.get("Remove", ""), "Interim": r.get("Interim", ""),
        "Gold_Canonical_Name": "", "Provenance": r.get("Provenance", "extracted"),
        "Possible_Duplicate_Appointment": r.get("Possible_Duplicate_Appointment", ""),
        "Unified_Attorney_ID": att_id(fn, ln) if (fn or ln) else "",
        "Unified_Firm_ID": firm_id(firm),
    }
    appts.append(a)

for a in appts:
    is_ind = bool(a["First Name"] or a["Last Name"]) and a["Appointee Type"] != "Firm"
    if is_ind:
        ind_rows += 1
        if a["Unified_Attorney_ID"]:
            linked_att += 1
    if a["Firm"] and a["Unified_Firm_ID"]:
        linked_firm += 1

# ===== 6. ATTORNEYS tab (roster + freshest demographics from the live cache) =====
att_tab = load_csv(ATT_ROSTER)
demo = {}
if os.path.exists(DEMO_CACHE):
    for line in open(DEMO_CACHE, encoding="utf-8"):
        if line.strip():
            try:                       # job may be appending concurrently; skip a partial last line
                d = json.loads(line)
                demo[d.get("Attorney_ID")] = d
            except json.JSONDecodeError:
                pass
demo_filled = 0
for r in att_tab:
    d = demo.get(r["Attorney_ID"])
    if d:
        r.update(AD.to_row_updates(d))
        if r.get("Sources"):
            demo_filled += 1
att_cols = list(att_tab[0].keys()) if att_tab else []

# ===== 7. FIRMS + MDLs =====
firm_tab = load_csv(FIRM_ROSTER)
firm_cols = list(firm_tab[0].keys()) if firm_tab else []
mdl_tab = load_csv(MDL_MASTER)
mdl_cols = list(mdl_tab[0].keys()) if mdl_tab else []

# ===== 8. write workbook =====
wb = openpyxl.Workbook()
def sheet(name, rows, cols):
    ws = wb.create_sheet(name)
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
wb.remove(wb.active)
sheet("MDLs", mdl_tab, mdl_cols)
sheet("Orders", orders, ORDER_COLS)
sheet("Appointments", appts, APPT_COLS)
sheet("Attorneys", att_tab, att_cols)
sheet("Firms", firm_tab, firm_cols)
wb.save(OUT)

print("\n=== UNIFIED DATABASE WRITTEN ===", os.path.relpath(OUT, ROOT))
print(f"  MDLs        : {len(mdl_tab):,}")
print(f"  Orders      : {len(orders):,}  (gold {sum(1 for o in orders if o['Corpus']=='old_gold')} + new {sum(1 for o in orders if o['Corpus']=='new_extracted')})")
print(f"  Appointments: {len(appts):,}  (gold {sum(1 for a in appts if a['Corpus']=='old_gold')} + new {sum(1 for a in appts if a['Corpus']=='new_extracted')})")
print(f"  Attorneys   : {len(att_tab):,}  (demographics filled so far: {demo_filled:,})")
print(f"  Firms       : {len(firm_tab):,}")
print(f"\n  LINKAGE: individual appt rows {ind_rows:,} -> Attorney_ID linked {linked_att:,} ({100*linked_att//max(1,ind_rows)}%)")
print(f"           appt rows with a firm -> Firm_ID linked {linked_firm:,}")
