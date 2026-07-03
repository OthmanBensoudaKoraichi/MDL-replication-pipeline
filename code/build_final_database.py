#!/usr/bin/env python3
"""Assemble the FINAL all-extracted unified database (2026-07-03 decision):
  primary tables = LLM extraction (old + new), deduped with v2; gold kept as reference tabs.
Tabs: MDLs | Orders | Appointments (+Unified_Attorney_ID/Firm_ID/Corpus) | Attorneys(v2+demographics)
      | Firms(v2) | Gold_Appointments | Gold_Attorneys | Role_Comparison(disagreement flag)
Linkage uses the exact dedup-v2 keys via the emitted maps."""
import csv, json, os, re, sys
import openpyxl
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "code"))
import dedup_v2 as V   # letters(), last_key(), firm_tokens

EXTRACT_XLSX = os.path.join(ROOT, "order_extractions.xlsx")
OUT = os.path.join(ROOT, "unified_mdl_database.xlsx")

def load_csv(p, enc="utf-8-sig"):
    with open(p, encoding=enc, newline="") as f:
        return list(csv.DictReader(f))

def read_tab(sheet):
    wb = openpyxl.load_workbook(EXTRACT_XLSX, read_only=True, data_only=True)
    ws = wb[sheet]; it = ws.iter_rows(values_only=True)
    hdr = [str(c) if c is not None else "" for c in next(it)]
    return hdr, [dict(zip(hdr, ["" if v is None else v for v in row])) for row in it]

def nmdl(x): return re.sub(r"^0+(\d)", r"\1", str(x or "").strip())

# corpus lookup
src = {}
for r in load_csv(os.path.join(ROOT, "MDL_merged.csv")):
    src[nmdl(r.get("MDL_NO"))] = (r.get("Source_Table") or "").strip()
def corpus(mdl):
    s = src.get(nmdl(mdl), "?")
    return {"old": "old", "both": "old", "new": "new"}.get(s, "?")

# maps
amap = {(r["first"], r["last"], r["mdl"]): r["Attorney_ID"] for r in load_csv(os.path.join(ROOT, "dedup_v2_attorney_map.csv"), "utf-8")}
fmap = {r["firm_key"]: r["Firm_ID"] for r in load_csv(os.path.join(ROOT, "dedup_v2_firm_map.csv"), "utf-8")}

def att_id(fn, ln, mdl):
    f = V.letters(str(fn).split()[0]) if str(fn).split() else ""
    l = V.last_key(ln)
    if f and f == l: f = ""
    return amap.get((f, l, nmdl(mdl)), "")

def firm_id(firm):
    k = " ".join(sorted(V.firm_tokens(firm)))
    return fmap.get(k, "")

# ---- Orders ----
ohdr, orders = read_tab("Orders")
for o in orders:
    o["Corpus"] = corpus(o.get("MDL_No"))
ocols = ["Corpus"] + ohdr

# ---- Appointments (+ links) ----
ahdr, appts = read_tab("Appointments")
lnk_att = lnk_firm = ind = 0
for a in appts:
    mdl = a.get("MDL_No")
    a["Corpus"] = corpus(mdl)
    fn, ln = a.get("First Name", ""), a.get("Last Name", "")
    is_ind = bool(str(fn).strip() or str(ln).strip()) and str(a.get("Appointee Type")).strip() != "Firm"
    a["Unified_Attorney_ID"] = att_id(fn, ln, mdl) if is_ind else ""
    a["Unified_Firm_ID"] = firm_id(a.get("Firm", ""))
    if is_ind:
        ind += 1
        if a["Unified_Attorney_ID"]: lnk_att += 1
    if str(a.get("Firm", "")).strip() and a["Unified_Firm_ID"]: lnk_firm += 1
acols = ["Corpus", "Unified_Attorney_ID", "Unified_Firm_ID"] + ahdr

# ---- Attorneys (v2): prefer the completed demographics roster if present ----
_demo_csv = os.path.join(ROOT, "canonical_attorneys_v2_demographics.csv")
if os.path.exists(_demo_csv):
    att_v2 = load_csv(_demo_csv)
    attcols = list(att_v2[0].keys())
    filled_gold = sum(1 for r in att_v2 if (r.get("Sources") or "").strip())
    filled_web = 0
    _USE_DEMO_CSV = True
else:
    _USE_DEMO_CSV = False
    att_v2 = load_csv(os.path.join(ROOT, "canonical_attorneys_v2.csv"))
    filled_gold = filled_web = 0
def nname(s): return re.sub(r"[^a-z]", "", str(s or "").lower())
# gold demographics by normalized canonical name
gold_demo = {}
for r in load_csv(os.path.join(ROOT, "csvs_current_dataset", "Attorneys-Export view.csv")):
    k = nname(r.get("Canonical_Name"))
    if k and any((r.get(c) or "").strip() for c in ("Gender", "Law_School_Name", "Birth_Year", "Bar_States")):
        gold_demo.setdefault(k, r)
# prior web-research cache -> map by the roster it was built on (name via unified v1 roster)
web_demo = {}
cache_path = os.path.join(ROOT, "demographics_cache_unified.jsonl")
v1_id_name = {}
if os.path.exists(os.path.join(ROOT, "canonical_attorneys_unified.csv")):
    for r in load_csv(os.path.join(ROOT, "canonical_attorneys_unified.csv")):
        v1_id_name[r["Attorney_ID"]] = nname(r["Canonical_Name"])
if os.path.exists(cache_path):
    for line in open(cache_path, encoding="utf-8"):
        if line.strip():
            try: d = json.loads(line)
            except: continue
            k = v1_id_name.get(d.get("Attorney_ID"))
            if k and d.get("sources"): web_demo.setdefault(k, d)
DEMOF = ["Gender", "Birth_Year", "Undergrad_School", "Law_School_Name", "Bar_States", "Sources"]
for r in ([] if _USE_DEMO_CSV else att_v2):
    k = nname(r.get("Canonical_Name"))
    g = gold_demo.get(k)
    if g:
        for c in DEMOF:
            r[c] = g.get(c, "")
        r["Demo_Source"] = "gold"; filled_gold += 1
    elif web_demo.get(k):
        d = web_demo[k]
        r["Gender"] = d.get("gender", ""); r["Birth_Year"] = d.get("birth_year", "")
        r["Undergrad_School"] = d.get("undergrad_school", ""); r["Law_School_Name"] = d.get("law_school", "")
        bar = d.get("bar_states") or []
        r["Bar_States"] = ", ".join(x for x in (bar if isinstance(bar, list) else [bar]) if x)
        srcs = d.get("sources") or []
        r["Sources"] = " | ".join(x for x in (srcs if isinstance(srcs, list) else [srcs]) if x)
        r["Demo_Source"] = "web"; filled_web += 1
    else:
        for c in DEMOF: r.setdefault(c, "")
        r["Demo_Source"] = ""
if not _USE_DEMO_CSV:
    attcols = list(att_v2[0].keys())

firm_v2 = load_csv(os.path.join(ROOT, "canonical_firms_v2.csv"))
gold_appts = load_csv(os.path.join(ROOT, "csvs_current_dataset", "Appointments-Grid view.csv"))
gold_atts = load_csv(os.path.join(ROOT, "csvs_current_dataset", "Attorneys-Export view.csv"))
role_cmp = load_csv(os.path.join(ROOT, "appointment_type_comparison.csv"))
mdl_tab = load_csv(os.path.join(ROOT, "MDL_merged.csv"))

# ---- write workbook ----
wb = openpyxl.Workbook(); wb.remove(wb.active)
def sheet(name, rows, cols):
    ws = wb.create_sheet(name[:31]); ws.append(cols)
    for r in rows: ws.append([r.get(c, "") for c in cols])
sheet("MDLs", mdl_tab, list(mdl_tab[0].keys()))
sheet("Orders", orders, ocols)
sheet("Appointments", appts, acols)
sheet("Attorneys", att_v2, attcols)
sheet("Firms", firm_v2, list(firm_v2[0].keys()))
sheet("Gold_Appointments", gold_appts, list(gold_appts[0].keys()))
sheet("Gold_Attorneys", gold_atts, list(gold_atts[0].keys()))
sheet("Role_Comparison", role_cmp, list(role_cmp[0].keys()))
wb.save(OUT)

print("=== FINAL all-extracted unified_mdl_database.xlsx ===")
print(f"  MDLs               : {len(mdl_tab):,}")
print(f"  Orders             : {len(orders):,}  (old {sum(1 for o in orders if o['Corpus']=='old')} / new {sum(1 for o in orders if o['Corpus']=='new')})")
print(f"  Appointments       : {len(appts):,}")
print(f"  Attorneys (v2)     : {len(att_v2):,}  | demographics: gold {filled_gold}, web {filled_web}, none {len(att_v2)-filled_gold-filled_web}")
print(f"  Firms (v2)         : {len(firm_v2):,}")
print(f"  Gold_Appointments  : {len(gold_appts):,}  | Gold_Attorneys: {len(gold_atts):,}  | Role_Comparison: {len(role_cmp):,}")
print(f"  LINKAGE: individual appts {ind:,} -> attorney_id {lnk_att:,} ({100*lnk_att//max(1,ind)}%) | firm links {lnk_firm:,}")
